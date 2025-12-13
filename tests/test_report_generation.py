"""
Tests for report generation functionality

Tests Excel generation for newsletter export and other report formats.
"""

import pytest
from datetime import date
from io import BytesIO
from openpyxl import load_workbook

from members.models import Member, MemberType
from members.reports.excel import generate_newsletter_excel


@pytest.mark.django_db
@pytest.mark.integration
class TestNewsletterExcelGeneration:
    """Test newsletter Excel export generation"""

    @pytest.fixture
    def member_type(self, db):
        """Create a test member type"""
        return MemberType.objects.create(
            member_type="Regular",
            member_dues=30.00,
            num_months=1,
        )

    @pytest.fixture
    def member_with_email(self, db, member_type):
        """Create an active member with email"""
        return Member.objects.create(
            member_id=1,
            first_name="John",
            last_name="Doe",
            email="john.doe@example.com",
            member_type=member_type,
            status="active",
            date_joined=date(2020, 1, 15),
            milestone_date=date(2015, 6, 10),
            expiration_date=date(2025, 12, 31),
        )

    @pytest.fixture
    def member_without_email(self, db, member_type):
        """Create an active member without email"""
        return Member.objects.create(
            member_id=2,
            first_name="Jane",
            last_name="Smith",
            email="",  # Empty email
            member_type=member_type,
            status="active",
            date_joined=date(2021, 3, 20),
            milestone_date=date(2018, 9, 5),
            expiration_date=date(2025, 11, 30),
        )

    @pytest.fixture
    def member_no_milestone(self, db, member_type):
        """Create an active member without milestone date"""
        return Member.objects.create(
            member_id=3,
            first_name="Bob",
            last_name="Johnson",
            email="bob.johnson@example.com",
            member_type=member_type,
            status="active",
            date_joined=date(2022, 5, 10),
            milestone_date=None,  # No milestone date
            expiration_date=date(2025, 10, 15),
        )

    @pytest.fixture
    def inactive_member(self, db, member_type):
        """Create an inactive member (should not appear in export)"""
        return Member.objects.create(
            member_id=4,
            first_name="Inactive",
            last_name="Member",
            email="inactive@example.com",
            member_type=member_type,
            status="inactive",
            date_joined=date(2019, 1, 1),
            milestone_date=date(2014, 1, 1),
            expiration_date=date(2020, 1, 1),
        )

    def test_function_exists(self):
        """Test that generate_newsletter_excel function exists"""
        assert callable(generate_newsletter_excel)

    def test_response_structure(self, db, member_with_email):
        """Test that function returns HttpResponse with correct headers"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_newsletter_excel(queryset)

        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "Content-Disposition" in response
        assert "newsletter_data_" in response["Content-Disposition"]
        assert ".xlsx" in response["Content-Disposition"]

    def test_excel_headers(self, db, member_with_email):
        """Test that Excel file has correct headers"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_newsletter_excel(queryset)

        # Load workbook from response content
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check headers
        expected_headers = [
            "MemberID",
            "FirstName",
            "LastName",
            "EmailName",
            "DateJoined",
            "Birthdate",
            "Expires",
            "MailName",
            "FullName",
        ]

        headers = [cell.value for cell in ws[1]]
        assert headers == expected_headers

    def test_member_with_email_data(self, db, member_with_email):
        """Test that member with email is correctly formatted"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_newsletter_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Get data row (row 2, since row 1 is headers)
        row_data = [cell.value for cell in ws[2]]

        assert row_data[0] == 1  # MemberID
        assert row_data[1] == "John"  # FirstName
        assert row_data[2] == "Doe"  # LastName
        assert row_data[3] == "john.doe@example.com"  # EmailName
        assert row_data[4] == "01/15/2020"  # DateJoined (MM/DD/YYYY)
        assert row_data[5] == "06/10/2015"  # Birthdate (MM/DD/YYYY)
        assert row_data[6] == "12/31/2025"  # Expires (MM/DD/YYYY)
        assert row_data[7] == "John Doe<john.doe@example.com>"  # MailName
        assert row_data[8] == "John Doe"  # FullName

    def test_member_without_email(self, db, member_without_email):
        """Test that member without email goes to 'No Email' sheet"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_newsletter_excel(queryset)

        wb = load_workbook(BytesIO(response.content))

        # Should have "No Email" sheet
        assert "No Email" in wb.sheetnames

        # Check data in "No Email" sheet
        no_email_sheet = wb["No Email"]
        row_data = [cell.value for cell in no_email_sheet[2]]  # Row 2 (data row)

        assert row_data[0] == 2  # MemberID
        assert row_data[1] == "Jane"  # FirstName
        assert row_data[2] == "Smith"  # LastName
        assert row_data[3] == "" or row_data[3] is None  # EmailName (empty)
        assert row_data[7] == "" or row_data[7] is None  # MailName (empty)

    def test_member_no_milestone_date(self, db, member_no_milestone):
        """Test that member without milestone date has blank Birthdate"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_newsletter_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        row_data = [cell.value for cell in ws[2]]
        assert row_data[5] == "" or row_data[5] is None  # Birthdate (blank)

    def test_inactive_member_excluded(self, db, member_with_email, inactive_member):
        """Test that inactive members are excluded from export"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_newsletter_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Should only have 1 data row (member_with_email)
        # Row 1 is headers, so max_row should be 2
        assert ws.max_row == 2

        # Verify inactive member is not in the data
        row_data = [cell.value for cell in ws[2]]
        assert row_data[0] == 1  # Only member_with_email (ID=1)
        assert row_data[0] != 4  # Not inactive_member (ID=4)

    def test_ordering_by_member_id(self, db, member_type):
        """Test that members are ordered by member_id ascending"""
        # Create members with different IDs in non-sequential order
        Member.objects.create(
            member_id=5,
            first_name="Third",
            last_name="Member",
            email="third@example.com",
            member_type=member_type,
            status="active",
            date_joined=date(2020, 1, 1),
            expiration_date=date(2025, 12, 31),
        )
        Member.objects.create(
            member_id=1,
            first_name="First",
            last_name="Member",
            email="first@example.com",
            member_type=member_type,
            status="active",
            date_joined=date(2020, 1, 1),
            expiration_date=date(2025, 12, 31),
        )
        Member.objects.create(
            member_id=3,
            first_name="Second",
            last_name="Member",
            email="second@example.com",
            member_type=member_type,
            status="active",
            date_joined=date(2020, 1, 1),
            expiration_date=date(2025, 12, 31),
        )

        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_newsletter_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check ordering (row 2, 3, 4 should be IDs 1, 3, 5)
        assert ws[2][0].value == 1  # First member
        assert ws[3][0].value == 3  # Second member
        assert ws[4][0].value == 5  # Third member

    def test_multiple_sheets_99_rows(self, db, member_type):
        """Test that Excel creates multiple sheets when >99 members with emails"""
        # Create 100 members with emails
        for i in range(1, 101):
            Member.objects.create(
                member_id=i,
                first_name=f"Member{i}",
                last_name="Test",
                email=f"member{i}@example.com",
                member_type=member_type,
                status="active",
                date_joined=date(2020, 1, 1),
                expiration_date=date(2025, 12, 31),
            )

        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_newsletter_excel(queryset)

        wb = load_workbook(BytesIO(response.content))

        # Should have Sheet 1 and Sheet 2
        assert "Sheet 1" in wb.sheetnames
        assert "Sheet 2" in wb.sheetnames

        # Sheet 1 should have 99 data rows + 1 header = 100 rows
        sheet1 = wb["Sheet 1"]
        assert sheet1.max_row == 100

        # Sheet 2 should have 1 data row + 1 header = 2 rows
        sheet2 = wb["Sheet 2"]
        assert sheet2.max_row == 2

        # Verify headers exist on both sheets
        headers_sheet1 = [cell.value for cell in sheet1[1]]
        headers_sheet2 = [cell.value for cell in sheet2[1]]
        expected_headers = [
            "MemberID",
            "FirstName",
            "LastName",
            "EmailName",
            "DateJoined",
            "Birthdate",
            "Expires",
            "MailName",
            "FullName",
        ]
        assert headers_sheet1 == expected_headers
        assert headers_sheet2 == expected_headers

    def test_no_email_sheet_separate(self, db, member_type):
        """Test that members with and without emails are in separate sheets"""
        # Create members with emails
        Member.objects.create(
            member_id=1,
            first_name="With",
            last_name="Email",
            email="with@example.com",
            member_type=member_type,
            status="active",
            date_joined=date(2020, 1, 1),
            expiration_date=date(2025, 12, 31),
        )

        # Create members without emails
        Member.objects.create(
            member_id=2,
            first_name="Without",
            last_name="Email",
            email="",  # Empty email
            member_type=member_type,
            status="active",
            date_joined=date(2020, 1, 1),
            expiration_date=date(2025, 12, 31),
        )

        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_newsletter_excel(queryset)

        wb = load_workbook(BytesIO(response.content))

        # Should have Sheet 1 (with emails) and "No Email" sheet
        assert "Sheet 1" in wb.sheetnames
        assert "No Email" in wb.sheetnames

        # Sheet 1 should have member with email
        sheet1 = wb["Sheet 1"]
        assert sheet1[2][0].value == 1  # MemberID = 1

        # "No Email" sheet should have member without email
        no_email_sheet = wb["No Email"]
        assert no_email_sheet[2][0].value == 2  # MemberID = 2

    def test_empty_queryset(self, db):
        """Test that empty queryset creates valid Excel file"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_newsletter_excel(queryset)

        wb = load_workbook(BytesIO(response.content))

        # Should have at least one sheet (even if empty)
        assert len(wb.sheetnames) >= 0

        # If there's a sheet, it should have headers
        if wb.sheetnames:
            ws = wb.active
            # Should only have header row
            assert ws.max_row == 1

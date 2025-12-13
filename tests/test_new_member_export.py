"""
Tests for New Member Export functionality

Tests the new member export view and Excel generation:
- new_member_export_view() GET and POST
- generate_new_member_excel() function
- Date range validation
- Member filtering by date_joined
"""

import pytest
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from django.test import Client
from django.contrib.auth.models import User
from openpyxl import load_workbook

from members.models import Member, MemberType
from members.reports.excel import generate_new_member_excel


@pytest.mark.django_db
@pytest.mark.integration
class TestNewMemberExportView:
    """Test new_member_export_view reports view"""

    @pytest.fixture
    def user(self, db):
        """Create a user for authentication"""
        return User.objects.create_user(
            username="testuser",
            password="testpass",
        )

    @pytest.fixture
    def client(self, user):
        """Create authenticated client"""
        client = Client()
        client.login(username="testuser", password="testpass")
        return client

    @pytest.fixture
    def member_type(self, db):
        """Create a test member type"""
        return MemberType.objects.create(
            member_type="Regular",
            member_dues=Decimal("30.00"),
            num_months=1,
        )

    def test_get_view_displays_form(self, client):
        """Test that GET request displays date range selection form"""
        response = client.get("/reports/new-members/")
        assert response.status_code == 200
        assert "New Member Export" in response.content.decode()
        assert "Start Date" in response.content.decode()
        assert "End Date" in response.content.decode()

    def test_get_view_sets_end_date_to_today(self, client):
        """Test that end_date field is pre-populated with today's date"""
        response = client.get("/reports/new-members/")
        assert response.status_code == 200
        content = response.content.decode()
        today_str = date.today().strftime("%Y-%m-%d")
        # Check that today's date appears in the end_date input value
        assert f'value="{today_str}"' in content or f'value="{today_str}' in content

    def test_get_view_requires_authentication(self, db):
        """Test that unauthenticated users cannot access the view"""
        client = Client()
        response = client.get("/reports/new-members/")
        # Should redirect to login
        assert response.status_code == 302

    def test_post_view_with_valid_dates_generates_excel(self, client, member_type):
        """Test that POST with valid dates generates Excel file"""
        # Create member within date range
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()

        Member.objects.create(
            first_name="John",
            last_name="Doe",
            member_type=member_type,
            status="active",
            date_joined=start_date + timedelta(days=10),
            expiration_date=date.today() + timedelta(days=30),
        )

        response = client.post(
            "/reports/new-members/",
            {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
        )

        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert ".xlsx" in response["Content-Disposition"]

    def test_post_view_validates_start_date_not_more_than_6_months_ago(self, client):
        """Test that start date cannot be more than 6 months ago"""
        start_date = date.today() - timedelta(days=185)  # More than 6 months
        end_date = date.today()

        response = client.post(
            "/reports/new-members/",
            {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
        )

        assert response.status_code == 200
        assert "6 months ago" in response.content.decode()

    def test_post_view_validates_end_date_not_in_future(self, client):
        """Test that end date cannot exceed today"""
        start_date = date.today() - timedelta(days=30)
        end_date = date.today() + timedelta(days=1)  # Tomorrow

        response = client.post(
            "/reports/new-members/",
            {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
        )

        assert response.status_code == 200
        assert "cannot exceed today" in response.content.decode()

    def test_post_view_validates_end_date_not_before_start_date(self, client):
        """Test that end date must be on or after start date"""
        start_date = date.today() - timedelta(days=10)
        end_date = date.today() - timedelta(days=20)  # Before start date

        response = client.post(
            "/reports/new-members/",
            {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
        )

        assert response.status_code == 200
        assert "on or after start date" in response.content.decode()

    def test_post_view_filters_by_date_range(self, client, member_type):
        """Test that only members within date range are included"""
        start_date = date.today() - timedelta(days=60)
        end_date = date.today() - timedelta(days=30)

        # Member within range
        member_in_range = Member.objects.create(
            first_name="In",
            last_name="Range",
            member_type=member_type,
            status="active",
            date_joined=start_date + timedelta(days=10),
            expiration_date=date.today() + timedelta(days=30),
        )

        # Member before range
        Member.objects.create(
            first_name="Before",
            last_name="Range",
            member_type=member_type,
            status="active",
            date_joined=start_date - timedelta(days=10),
            expiration_date=date.today() + timedelta(days=30),
        )

        # Member after range
        Member.objects.create(
            first_name="After",
            last_name="Range",
            member_type=member_type,
            status="active",
            date_joined=end_date + timedelta(days=10),
            expiration_date=date.today() + timedelta(days=30),
        )

        response = client.post(
            "/reports/new-members/",
            {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
        )

        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check that only member_in_range is in the Excel file
        # Row 1 is headers, so check row 2
        if ws.max_row > 1:
            row_data = [cell.value for cell in ws[2]]
            assert row_data[1] == "In"  # FirstName
            assert row_data[2] == "Range"  # LastName

    def test_post_view_only_includes_active_members(self, client, member_type):
        """Test that only active members are included"""
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()

        # Active member
        Member.objects.create(
            first_name="Active",
            last_name="Member",
            member_type=member_type,
            status="active",
            date_joined=start_date + timedelta(days=10),
            expiration_date=date.today() + timedelta(days=30),
        )

        # Inactive member
        Member.objects.create(
            first_name="Inactive",
            last_name="Member",
            member_type=member_type,
            status="inactive",
            date_joined=start_date + timedelta(days=10),
            expiration_date=date.today() + timedelta(days=30),
        )

        response = client.post(
            "/reports/new-members/",
            {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
        )

        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check that only active member is included
        if ws.max_row > 1:
            row_data = [cell.value for cell in ws[2]]
            assert row_data[1] == "Active"  # FirstName
            assert "Inactive" not in [
                cell.value for row in ws.iter_rows(min_row=2) for cell in row
            ]


@pytest.mark.django_db
@pytest.mark.integration
class TestNewMemberExcelGeneration:
    """Test generate_new_member_excel Excel generation function"""

    @pytest.fixture
    def member_type(self, db):
        """Create a test member type"""
        return MemberType.objects.create(
            member_type="Regular",
            member_dues=Decimal("30.00"),
            num_months=1,
        )

    @pytest.fixture
    def member_with_email(self, db, member_type):
        """Create an active member with email and full address"""
        return Member.objects.create(
            member_id=1,
            first_name="John",
            last_name="Doe",
            email="john.doe@example.com",
            member_type=member_type,
            status="active",
            date_joined=date(2024, 1, 15),
            milestone_date=date(2015, 6, 10),
            expiration_date=date(2025, 12, 31),
            home_address="550 Kiely Blvd.",
            home_city="San Jose",
            home_state="CA",
            home_zip="95117",
            home_phone="(408) 555-1234",
        )

    @pytest.fixture
    def member_without_email(self, db, member_type):
        """Create an active member without email"""
        return Member.objects.create(
            member_id=2,
            first_name="Jane",
            last_name="Smith",
            email="",  # Empty email
            member_type=member_type,
            status="active",
            date_joined=date(2024, 2, 20),
            milestone_date=date(2018, 9, 5),
            expiration_date=date(2025, 11, 30),
            home_address="123 Main St",
            home_city="Los Angeles",
            home_state="CA",
            home_zip="90001",
            home_phone="(213) 555-5678",
        )

    def test_response_structure(self, db, member_with_email):
        """Test that response has correct structure"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_new_member_excel(queryset)

        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert ".xlsx" in response["Content-Disposition"]
        assert "new_members_" in response["Content-Disposition"]

    def test_excel_headers(self, db, member_with_email):
        """Test that Excel file has correct headers"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_new_member_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        expected_headers = [
            "MemberID",
            "FirstName",
            "LastName",
            "AddressLong",
            "HomeAddress",
            "HomeCity",
            "HomeState",
            "Zip5",
            "Phone",
            "EmailName",
            "Birthdate",
            "DateJoined",
            "Expires",
            "MailName",
        ]

        headers = [cell.value for cell in ws[1]]
        assert headers == expected_headers

    def test_address_long_format(self, db, member_with_email):
        """Test that AddressLong is formatted correctly: street[TAB]city, state[TAB]zip"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_new_member_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Get AddressLong value (column 4, row 2)
        address_long = ws.cell(row=2, column=4).value
        # Should be: "550 Kiely Blvd.\tSan Jose, CA\t95117"
        assert "550 Kiely Blvd." in address_long
        assert "San Jose, CA" in address_long
        assert "95117" in address_long
        # Check for tabs (should have 2 tabs)
        assert address_long.count("\t") == 2

    def test_separate_home_state_and_zip5_columns(self, db, member_with_email):
        """Test that HomeState and Zip5 are separate columns"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_new_member_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Row 2, columns 7 and 8 (HomeState and Zip5)
        home_state = ws.cell(row=2, column=7).value
        zip5 = ws.cell(row=2, column=8).value

        assert home_state == "CA"
        assert zip5 == "95117"

    def test_member_with_email_data(self, db, member_with_email):
        """Test that member with email is correctly formatted"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_new_member_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Get data row (row 2, since row 1 is headers)
        row_data = [cell.value for cell in ws[2]]

        assert row_data[0] == 1  # MemberID
        assert row_data[1] == "John"  # FirstName
        assert row_data[2] == "Doe"  # LastName
        assert row_data[4] == "550 Kiely Blvd."  # HomeAddress
        assert row_data[5] == "San Jose"  # HomeCity
        assert row_data[6] == "CA"  # HomeState
        assert row_data[7] == "95117"  # Zip5
        assert row_data[8] == "(408) 555-1234"  # Phone
        assert row_data[9] == "john.doe@example.com"  # EmailName
        assert row_data[10] == "06/10/2015"  # Birthdate (MM/DD/YYYY)
        assert row_data[11] == "01/15/2024"  # DateJoined (MM/DD/YYYY)
        assert row_data[12] == "12/31/2025"  # Expires (MM/DD/YYYY)
        assert row_data[13] == "John Doe<john.doe@example.com>"  # MailName

    def test_member_without_email_goes_to_no_email_sheet(
        self, db, member_without_email
    ):
        """Test that member without email goes to 'no email' sheet"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_new_member_excel(queryset)

        wb = load_workbook(BytesIO(response.content))

        # Should have "no email" sheet
        assert "no email" in wb.sheetnames

        no_email_sheet = wb["no email"]
        # Check that member is in no email sheet
        row_data = [cell.value for cell in no_email_sheet[2]]
        assert row_data[1] == "Jane"  # FirstName
        assert row_data[2] == "Smith"  # LastName
        assert (
            row_data[9] is None or row_data[9] == ""
        )  # EmailName should be empty (None in Excel)
        assert (
            row_data[13] is None or row_data[13] == ""
        )  # MailName should be empty (None in Excel)

    def test_member_with_email_goes_to_main_sheet(self, db, member_with_email):
        """Test that member with email goes to main sheet"""
        queryset = Member.objects.filter(status="active").order_by("member_id")
        response = generate_new_member_excel(queryset)

        wb = load_workbook(BytesIO(response.content))

        # Should have main sheet (first sheet)
        sheet_names = wb.sheetnames
        main_sheet_name = [name for name in sheet_names if name != "no email"][0]
        main_sheet = wb[main_sheet_name]

        # Check that member is in main sheet
        row_data = [cell.value for cell in main_sheet[2]]
        assert row_data[1] == "John"  # FirstName
        assert row_data[9] == "john.doe@example.com"  # EmailName

    def test_zip_extraction_first_5_digits(self, db, member_type):
        """Test that zip code is extracted as first 5 digits"""
        # Create member with zip code longer than 5 digits
        member = Member.objects.create(
            member_id=3,
            first_name="Test",
            last_name="Zip",
            email="test@example.com",
            member_type=member_type,
            status="active",
            date_joined=date(2024, 3, 1),
            expiration_date=date(2025, 12, 31),
            home_zip="95117-1234",  # 9-digit zip
        )

        queryset = Member.objects.filter(status="active", member_id=3)
        response = generate_new_member_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        zip5 = ws.cell(row=2, column=8).value
        assert zip5 == "95117"  # Should be first 5 digits only

    def test_empty_queryset_creates_empty_sheet(self, db):
        """Test that empty queryset creates sheet with headers only"""
        queryset = Member.objects.none()
        response = generate_new_member_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Should have headers but no data rows
        assert ws.max_row == 1
        headers = [cell.value for cell in ws[1]]
        assert len(headers) == 14  # Should have 14 columns

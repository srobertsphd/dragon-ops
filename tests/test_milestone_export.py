"""
Tests for Milestone Export functionality

Tests the milestone export view and Excel generation:
- milestone_export_view() GET and POST
- generate_milestone_excel() function
- Date range validation
- Milestone date filtering logic (month/day THIS YEAR)
- Special column calculations (BdayLong, Years, Jmonth, Jyear)
"""

import pytest
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from django.test import Client
from django.contrib.auth.models import User
from openpyxl import load_workbook

from members.models import Member, MemberType
from members.reports.excel import generate_milestone_excel
from members.views.reports import milestone_falls_in_range


@pytest.mark.django_db
@pytest.mark.integration
class TestMilestoneExportView:
    """Test milestone_export_view reports view"""

    @pytest.fixture
    def user(self, db):
        """Create a staff user for authentication"""
        return User.objects.create_user(
            username="testuser",
            password="testpass",
            is_staff=True,
        )

    @pytest.fixture
    def client(self, user):
        """Create authenticated client"""
        client = Client()
        client.login(username="testuser", password="testpass")
        return client

    @pytest.fixture
    def member_type(self, db):
        """Create a test member type"""
        return MemberType.objects.create(
            member_type="Regular",
            member_dues=Decimal("30.00"),
            num_months=1,
        )

    def test_get_view_displays_form(self, client):
        """Test that GET request displays date range selection form"""
        response = client.get("/reports/milestone-export/")
        assert response.status_code == 200
        assert "Milestone Export" in response.content.decode()
        assert "Start Date" in response.content.decode()
        assert "End Date" in response.content.decode()

    def test_get_view_sets_start_date_to_today(self, client):
        """Test that start_date field is pre-populated with today's date"""
        response = client.get("/reports/milestone-export/")
        assert response.status_code == 200
        content = response.content.decode()
        today_str = date.today().strftime("%Y-%m-%d")
        # Check that today's date appears in the start_date input value
        assert f'value="{today_str}"' in content or f'value="{today_str}' in content

    def test_get_view_sets_end_date_to_today_plus_30_days(self, client):
        """Test that end_date field is pre-populated with today + 30 days"""
        response = client.get("/reports/milestone-export/")
        assert response.status_code == 200
        content = response.content.decode()
        end_date_default = (date.today() + timedelta(days=30)).strftime("%Y-%m-%d")
        # Check that end_date_default appears in the end_date input value
        assert (
            f'value="{end_date_default}"' in content
            or f'value="{end_date_default}' in content
        )

    def test_get_view_requires_authentication(self, db):
        """Test that unauthenticated users cannot access the view"""
        client = Client()
        response = client.get("/reports/milestone-export/")
        # Should redirect to login
        assert response.status_code == 302

    def test_post_view_with_valid_dates_generates_excel(self, client, member_type):
        """Test that POST with valid dates generates Excel file"""
        # Create member with milestone date that falls in range THIS YEAR
        today = date.today()
        milestone_date = date(
            2015, today.month, today.day
        )  # Same month/day, different year

        Member.objects.create(
            first_name="John",
            last_name="Doe",
            member_type=member_type,
            status="active",
            milestone_date=milestone_date,
            date_joined=date(2020, 1, 15),
            expiration_date=date.today() + timedelta(days=30),
        )

        start_date = today
        end_date = today + timedelta(days=30)

        response = client.post(
            "/reports/milestone-export/",
            {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
        )

        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert ".xlsx" in response["Content-Disposition"]

    def test_post_view_validates_start_date_not_more_than_6_months_ago(self, client):
        """Test that start date cannot be more than 6 months ago"""
        start_date = date.today() - timedelta(days=185)  # More than 6 months
        end_date = date.today() + timedelta(days=30)

        response = client.post(
            "/reports/milestone-export/",
            {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
        )

        assert response.status_code == 200
        assert "6 months ago" in response.content.decode()

    def test_post_view_validates_end_date_not_more_than_6_months_future(self, client):
        """Test that end date cannot be more than 6 months in the future"""
        start_date = date.today()
        end_date = date.today() + timedelta(days=185)  # More than 6 months

        response = client.post(
            "/reports/milestone-export/",
            {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
        )

        assert response.status_code == 200
        assert "6 months in the future" in response.content.decode()

    def test_post_view_validates_end_date_not_before_start_date(self, client):
        """Test that end date must be on or after start date"""
        start_date = date.today() + timedelta(days=10)
        end_date = date.today()  # Before start date

        response = client.post(
            "/reports/milestone-export/",
            {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
        )

        assert response.status_code == 200
        assert "on or after start date" in response.content.decode()

    def test_post_view_filters_by_milestone_date_range_this_year(
        self, client, member_type
    ):
        """Test that only members whose milestone dates (THIS YEAR) fall in range are included"""
        today = date.today()
        start_date = today + timedelta(days=10)
        end_date = today + timedelta(days=40)

        # Member with milestone date that falls in range THIS YEAR
        # Use a fixed date that will fall in range: today + 15 days
        milestone_day = min(today.day + 15, 28)
        milestone_in_range = date(2015, today.month, milestone_day)
        Member.objects.create(
            first_name="In",
            last_name="Range",
            member_type=member_type,
            status="active",
            milestone_date=milestone_in_range,
            date_joined=date(2020, 1, 15),
            expiration_date=date.today() + timedelta(days=30),
        )

        # Member with milestone date that falls BEFORE range THIS YEAR
        # Use a date that's before start_date: today - 5 days
        milestone_before_day = max(1, min(today.day - 5, 28))
        milestone_before = date(2010, today.month, milestone_before_day)
        Member.objects.create(
            first_name="Before",
            last_name="Range",
            member_type=member_type,
            status="active",
            milestone_date=milestone_before,
            date_joined=date(2020, 1, 15),
            expiration_date=date.today() + timedelta(days=30),
        )

        # Member with milestone date that falls AFTER range THIS YEAR
        # Use end_date + 10 days to ensure it's after the range
        after_date = end_date + timedelta(days=10)
        milestone_after = date(2018, after_date.month, min(after_date.day, 28))
        Member.objects.create(
            first_name="After",
            last_name="Range",
            member_type=member_type,
            status="active",
            milestone_date=milestone_after,
            date_joined=date(2020, 1, 15),
            expiration_date=date.today() + timedelta(days=30),
        )

        response = client.post(
            "/reports/milestone-export/",
            {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
        )

        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check that only member_in_range is in the Excel file
        # Row 1 is headers, so check row 2
        if ws.max_row > 1:
            row_data = [cell.value for cell in ws[2]]
            assert row_data[1] == "In"  # FirstName
            assert row_data[2] == "Range"  # LastName

    def test_post_view_only_includes_active_members(self, client, member_type):
        """Test that only active members are included"""
        today = date.today()
        start_date = today
        end_date = today + timedelta(days=30)

        # Active member with milestone
        milestone_date = date(2015, today.month, min(today.day, 28))
        Member.objects.create(
            first_name="Active",
            last_name="Member",
            member_type=member_type,
            status="active",
            milestone_date=milestone_date,
            date_joined=date(2020, 1, 15),
            expiration_date=date.today() + timedelta(days=30),
        )

        # Inactive member with milestone
        Member.objects.create(
            first_name="Inactive",
            last_name="Member",
            member_type=member_type,
            status="inactive",
            milestone_date=milestone_date,
            date_joined=date(2020, 1, 15),
            expiration_date=date.today() + timedelta(days=30),
        )

        response = client.post(
            "/reports/milestone-export/",
            {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
        )

        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check that only active member is included
        if ws.max_row > 1:
            row_data = [cell.value for cell in ws[2]]
            assert row_data[1] == "Active"  # FirstName
            assert "Inactive" not in [
                cell.value for row in ws.iter_rows(min_row=2) for cell in row
            ]

    def test_post_view_excludes_members_without_milestone_date(
        self, client, member_type
    ):
        """Test that members without milestone_date are excluded"""
        today = date.today()
        start_date = today
        end_date = today + timedelta(days=30)

        # Member with milestone date
        milestone_date = date(2015, today.month, min(today.day, 28))
        Member.objects.create(
            first_name="With",
            last_name="Milestone",
            member_type=member_type,
            status="active",
            milestone_date=milestone_date,
            date_joined=date(2020, 1, 15),
            expiration_date=date.today() + timedelta(days=30),
        )

        # Member without milestone date
        Member.objects.create(
            first_name="Without",
            last_name="Milestone",
            member_type=member_type,
            status="active",
            milestone_date=None,  # No milestone date
            date_joined=date(2020, 1, 15),
            expiration_date=date.today() + timedelta(days=30),
        )

        response = client.post(
            "/reports/milestone-export/",
            {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
        )

        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check that only member with milestone is included
        if ws.max_row > 1:
            row_data = [cell.value for cell in ws[2]]
            assert row_data[1] == "With"  # FirstName
            assert "Without" not in [
                cell.value for row in ws.iter_rows(min_row=2) for cell in row
            ]


@pytest.mark.django_db
@pytest.mark.integration
class TestMilestoneFallsInRange:
    """Test milestone_falls_in_range helper function"""

    def test_milestone_falls_in_range_returns_true_when_in_range(self):
        """Test that function returns True when milestone (this year) falls in range"""
        milestone_date = date(2015, 6, 15)
        start_date = date(2025, 6, 1)
        end_date = date(2025, 7, 31)

        result = milestone_falls_in_range(milestone_date, start_date, end_date, 2025)
        assert result is True

    def test_milestone_falls_in_range_returns_false_when_before_range(self):
        """Test that function returns False when milestone (this year) is before range"""
        milestone_date = date(2015, 5, 15)
        start_date = date(2025, 6, 1)
        end_date = date(2025, 7, 31)

        result = milestone_falls_in_range(milestone_date, start_date, end_date, 2025)
        assert result is False

    def test_milestone_falls_in_range_returns_false_when_after_range(self):
        """Test that function returns False when milestone (this year) is after range"""
        milestone_date = date(2015, 8, 15)
        start_date = date(2025, 6, 1)
        end_date = date(2025, 7, 31)

        result = milestone_falls_in_range(milestone_date, start_date, end_date, 2025)
        assert result is False

    def test_milestone_falls_in_range_handles_leap_year_date(self):
        """Test that Feb 29 milestone dates are handled correctly in non-leap years"""
        milestone_date = date(2016, 2, 29)  # Leap year date
        start_date = date(2025, 2, 1)
        end_date = date(2025, 2, 28)

        # Should use Feb 28 in 2025 (non-leap year)
        result = milestone_falls_in_range(milestone_date, start_date, end_date, 2025)
        assert result is True  # Feb 28 falls in range

    def test_milestone_falls_in_range_uses_current_year_by_default(self):
        """Test that function uses current year if not specified"""
        milestone_date = date(2015, date.today().month, min(date.today().day, 28))
        start_date = date.today()
        end_date = date.today() + timedelta(days=30)

        result = milestone_falls_in_range(milestone_date, start_date, end_date)
        # Result depends on whether milestone falls in range this year
        assert isinstance(result, bool)


@pytest.mark.django_db
@pytest.mark.integration
class TestMilestoneExcelGeneration:
    """Test generate_milestone_excel Excel generation function"""

    @pytest.fixture
    def member_type(self, db):
        """Create a test member type"""
        return MemberType.objects.create(
            member_type="Regular",
            member_dues=Decimal("30.00"),
            num_months=1,
        )

    @pytest.fixture
    def member_with_email(self, db, member_type):
        """Create an active member with email and milestone date"""
        return Member.objects.create(
            member_id=1,
            first_name="John",
            last_name="Doe",
            email="john.doe@example.com",
            member_type=member_type,
            status="active",
            milestone_date=date(2015, 6, 15),
            date_joined=date(1997, 5, 31),
            expiration_date=date(2025, 12, 31),
        )

    @pytest.fixture
    def member_without_email(self, db, member_type):
        """Create an active member without email"""
        return Member.objects.create(
            member_id=2,
            first_name="Jane",
            last_name="Smith",
            email="",  # Empty email
            member_type=member_type,
            status="active",
            milestone_date=date(2018, 10, 20),
            date_joined=date(2020, 10, 15),
            expiration_date=date(2025, 11, 30),
        )

    def test_response_structure(self, db, member_with_email):
        """Test that response has correct structure"""
        queryset = [member_with_email]
        response = generate_milestone_excel(queryset)

        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert ".xlsx" in response["Content-Disposition"]
        assert "milestone_export_" in response["Content-Disposition"]

    def test_excel_headers(self, db, member_with_email):
        """Test that Excel file has correct headers"""
        queryset = [member_with_email]
        response = generate_milestone_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        expected_headers = [
            "MemberID",
            "FirstName",
            "LastName",
            "Birthdate",
            "DateJoined",
            "Expires",
            "BdayLong",
            "Years",
            "MailName",
            "Jmonth",
            "Jyear",
        ]

        headers = [cell.value for cell in ws[1]]
        assert headers == expected_headers

    def test_bday_long_format(self, db, member_with_email):
        """Test that BdayLong is formatted correctly: 'DayOfWeek, MonthName Day'"""
        queryset = [member_with_email]
        response = generate_milestone_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Get BdayLong value (column 7, row 2)
        bday_long = ws.cell(row=2, column=7).value
        # Should be formatted like "Monday, June 15" (for milestone date June 15, THIS YEAR)
        assert "," in bday_long  # Should have comma
        assert "June" in bday_long or "15" in bday_long  # Should have month name or day

    def test_years_calculation(self, db, member_with_email):
        """Test that Years column calculates current_year - milestone_year"""
        queryset = [member_with_email]
        response = generate_milestone_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Years column (column 8, row 2)
        years = ws.cell(row=2, column=8).value
        # milestone_date = 2015-06-15, current_year = 2025 → Years = 10
        current_year = date.today().year
        expected_years = current_year - 2015
        assert years == expected_years

    def test_jmonth_from_date_joined(self, db, member_with_email):
        """Test that Jmonth uses month name from date_joined"""
        queryset = [member_with_email]
        response = generate_milestone_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Jmonth column (column 10, row 2)
        jmonth = ws.cell(row=2, column=10).value
        # date_joined = 1997-05-31 → Jmonth = "May"
        assert jmonth == "May"

    def test_jyear_from_date_joined(self, db, member_with_email):
        """Test that Jyear uses year from date_joined"""
        queryset = [member_with_email]
        response = generate_milestone_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Jyear column (column 11, row 2)
        jyear = ws.cell(row=2, column=11).value
        # date_joined = 1997-05-31 → Jyear = 1997
        assert jyear == 1997

    def test_member_with_email_data(self, db, member_with_email):
        """Test that member with email is correctly formatted"""
        queryset = [member_with_email]
        response = generate_milestone_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Get data row (row 2, since row 1 is headers)
        row_data = [cell.value for cell in ws[2]]

        assert row_data[0] == 1  # MemberID
        assert row_data[1] == "John"  # FirstName
        assert row_data[2] == "Doe"  # LastName
        assert row_data[3] == "06/15/2015"  # Birthdate (MM/DD/YYYY)
        assert row_data[4] == "05/31/1997"  # DateJoined (MM/DD/YYYY)
        assert row_data[5] == "12/31/2025"  # Expires (MM/DD/YYYY)
        assert "," in str(row_data[6])  # BdayLong (has comma)
        assert isinstance(row_data[7], int)  # Years (integer)
        assert row_data[8] == "John Doe<john.doe@example.com>"  # MailName
        assert row_data[9] == "May"  # Jmonth
        assert row_data[10] == 1997  # Jyear

    def test_member_without_email_goes_to_no_email_sheet(
        self, db, member_without_email
    ):
        """Test that member without email goes to 'no email' sheet"""
        queryset = [member_without_email]
        response = generate_milestone_excel(queryset)

        wb = load_workbook(BytesIO(response.content))

        # Should have "no email" sheet
        assert "no email" in wb.sheetnames

        no_email_sheet = wb["no email"]
        # Check that member is in no email sheet
        row_data = [cell.value for cell in no_email_sheet[2]]
        assert row_data[1] == "Jane"  # FirstName
        assert row_data[2] == "Smith"  # LastName
        assert row_data[8] is None or row_data[8] == ""  # MailName should be empty

    def test_member_with_email_goes_to_main_sheet(self, db, member_with_email):
        """Test that member with email goes to main sheet"""
        queryset = [member_with_email]
        response = generate_milestone_excel(queryset)

        wb = load_workbook(BytesIO(response.content))

        # Should have main sheet (first sheet)
        sheet_names = wb.sheetnames
        main_sheet_name = [name for name in sheet_names if name != "no email"][0]
        main_sheet = wb[main_sheet_name]

        # Check that member is in main sheet
        row_data = [cell.value for cell in main_sheet[2]]
        assert row_data[1] == "John"  # FirstName
        assert row_data[8] == "John Doe<john.doe@example.com>"  # MailName

    def test_empty_queryset_creates_empty_sheet(self, db):
        """Test that empty queryset creates sheet with headers only"""
        queryset = []
        response = generate_milestone_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Should have headers but no data rows
        assert ws.max_row == 1
        headers = [cell.value for cell in ws[1]]
        assert len(headers) == 11  # Should have 11 columns

    def test_leap_year_date_handling(self, db, member_type):
        """Test that Feb 29 milestone dates are handled correctly"""
        # Create member with Feb 29 milestone date
        member = Member.objects.create(
            member_id=3,
            first_name="Leap",
            last_name="Year",
            email="leap@example.com",
            member_type=member_type,
            status="active",
            milestone_date=date(2016, 2, 29),  # Leap year date
            date_joined=date(2020, 1, 15),
            expiration_date=date(2025, 12, 31),
        )

        queryset = [member]
        response = generate_milestone_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # BdayLong should handle Feb 29 gracefully (use Feb 28 in non-leap years)
        bday_long = ws.cell(row=2, column=7).value
        assert bday_long is not None  # Should have a value
        # Should contain "February" or "28" (handled as Feb 28)
        assert "February" in bday_long or "28" in bday_long

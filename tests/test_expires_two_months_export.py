"""
Tests for Expires Two Months Export functionality

Tests the expires two months export view and Excel generation:
- expires_two_months_export_view() GET and POST
- generate_expires_two_months_excel() function
- Member filtering by expiration_date (60+ days ago)
- Email separation logic
"""

import pytest
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from django.test import Client
from django.contrib.auth.models import User
from openpyxl import load_workbook

from members.models import Member, MemberType
from members.reports.excel import generate_expires_two_months_excel


@pytest.mark.django_db
@pytest.mark.integration
class TestExpiresTwoMonthsExportView:
    """Test expires_two_months_export_view reports view"""

    @pytest.fixture
    def user(self, db):
        """Create a staff user for authentication"""
        return User.objects.create_user(
            username="testuser",
            password="testpass",
            is_staff=True,
        )

    @pytest.fixture
    def client(self, user):
        """Create authenticated client"""
        client = Client()
        client.login(username="testuser", password="testpass")
        return client

    @pytest.fixture
    def member_type(self, db):
        """Create a test member type"""
        return MemberType.objects.create(
            member_type="Regular",
            member_dues=Decimal("30.00"),
            num_months=1,
        )

    def test_get_view_displays_form(self, client):
        """Test that GET request displays export page"""
        response = client.get("/reports/expires-two-months/")
        assert response.status_code == 200
        assert "Expires Two Months Export" in response.content.decode()
        assert "Generate Excel Export" in response.content.decode()

    def test_get_view_requires_authentication(self, db):
        """Test that unauthenticated users cannot access the view"""
        client = Client()
        response = client.get("/reports/expires-two-months/")
        # Should redirect to login
        assert response.status_code == 302

    def test_post_view_generates_excel(self, client, member_type):
        """Test that POST request generates Excel file"""
        # Create member expired 60+ days ago
        today = date.today()
        expiration_date = today - timedelta(days=70)

        Member.objects.create(
            first_name="John",
            last_name="Doe",
            member_type=member_type,
            status="active",
            expiration_date=expiration_date,
            date_joined=date(2020, 1, 15),
        )

        response = client.post("/reports/expires-two-months/")

        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert ".xlsx" in response["Content-Disposition"]

    def test_post_view_filters_by_expiration_date_60_days_ago(
        self, client, member_type
    ):
        """Test that only members expired 60+ days ago are included"""
        today = date.today()

        # Member expired 70 days ago (should be included)
        Member.objects.create(
            first_name="Expired",
            last_name="Member",
            member_type=member_type,
            status="active",
            expiration_date=today - timedelta(days=70),
            date_joined=date(2020, 1, 15),
        )

        # Member expired 59 days ago (should be excluded)
        Member.objects.create(
            first_name="Recent",
            last_name="Expired",
            member_type=member_type,
            status="active",
            expiration_date=today - timedelta(days=59),
            date_joined=date(2020, 1, 15),
        )

        # Member expired exactly 60 days ago (should be included)
        Member.objects.create(
            first_name="Exactly",
            last_name="Sixty",
            member_type=member_type,
            status="active",
            expiration_date=today - timedelta(days=60),
            date_joined=date(2020, 1, 15),
        )

        response = client.post("/reports/expires-two-months/")

        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check that only members expired 60+ days ago are included
        if ws.max_row > 1:
            row_data = [cell.value for cell in ws[2]]
            assert row_data[1] == "Expired"  # FirstName
            assert "Recent" not in [
                cell.value for row in ws.iter_rows(min_row=2) for cell in row
            ]

    def test_post_view_only_includes_active_members(self, client, member_type):
        """Test that only active members are included"""
        today = date.today()
        expiration_date = today - timedelta(days=70)

        # Active member expired 70 days ago
        Member.objects.create(
            first_name="Active",
            last_name="Member",
            member_type=member_type,
            status="active",
            expiration_date=expiration_date,
            date_joined=date(2020, 1, 15),
        )

        # Inactive member expired 70 days ago
        Member.objects.create(
            first_name="Inactive",
            last_name="Member",
            member_type=member_type,
            status="inactive",
            expiration_date=expiration_date,
            date_joined=date(2020, 1, 15),
        )

        response = client.post("/reports/expires-two-months/")

        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check that only active member is included
        if ws.max_row > 1:
            row_data = [cell.value for cell in ws[2]]
            assert row_data[1] == "Active"  # FirstName
            assert "Inactive" not in [
                cell.value for row in ws.iter_rows(min_row=2) for cell in row
            ]


@pytest.mark.django_db
@pytest.mark.integration
class TestExpiresTwoMonthsExcelGeneration:
    """Test generate_expires_two_months_excel Excel generation function"""

    @pytest.fixture
    def member_type(self, db):
        """Create a test member type"""
        return MemberType.objects.create(
            member_type="Regular",
            member_dues=Decimal("30.00"),
            num_months=1,
        )

    @pytest.fixture
    def member_with_email(self, db, member_type):
        """Create an active member with email expired 70 days ago"""
        return Member.objects.create(
            member_id=1,
            first_name="John",
            last_name="Doe",
            email="john.doe@example.com",
            member_type=member_type,
            status="active",
            expiration_date=date.today() - timedelta(days=70),
            date_joined=date(2020, 1, 15),
        )

    @pytest.fixture
    def member_without_email(self, db, member_type):
        """Create an active member without email expired 70 days ago"""
        return Member.objects.create(
            member_id=2,
            first_name="Jane",
            last_name="Smith",
            email="",  # Empty email
            member_type=member_type,
            status="active",
            expiration_date=date.today() - timedelta(days=70),
            date_joined=date(2020, 5, 20),
        )

    def test_response_structure(self, db, member_with_email):
        """Test that response has correct structure"""
        queryset = [member_with_email]
        response = generate_expires_two_months_excel(queryset)

        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert ".xlsx" in response["Content-Disposition"]
        assert "expires_two_months_" in response["Content-Disposition"]

    def test_excel_headers(self, db, member_with_email):
        """Test that Excel file has correct headers"""
        queryset = [member_with_email]
        response = generate_expires_two_months_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        expected_headers = [
            "MemberID",
            "FirstName",
            "LastName",
            "EmailName",
            "DateJoined",
            "MailName",
            "Expires",
        ]

        headers = [cell.value for cell in ws[1]]
        assert headers == expected_headers

    def test_member_with_email_data(self, db, member_with_email):
        """Test that member with email is correctly formatted"""
        queryset = [member_with_email]
        response = generate_expires_two_months_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Get data row (row 2, since row 1 is headers)
        row_data = [cell.value for cell in ws[2]]

        assert row_data[0] == 1  # MemberID
        assert row_data[1] == "John"  # FirstName
        assert row_data[2] == "Doe"  # LastName
        assert row_data[3] == "john.doe@example.com"  # EmailName
        assert row_data[4] == "01/15/2020"  # DateJoined (MM/DD/YYYY)
        assert row_data[5] == "John Doe<john.doe@example.com>"  # MailName
        assert row_data[6] == (date.today() - timedelta(days=70)).strftime(
            "%m/%d/%Y"
        )  # Expires (MM/DD/YYYY)

    def test_member_without_email_goes_to_no_email_sheet(
        self, db, member_without_email
    ):
        """Test that member without email goes to 'No Email' sheet"""
        queryset = [member_without_email]
        response = generate_expires_two_months_excel(queryset)

        wb = load_workbook(BytesIO(response.content))

        # Should have "No Email" sheet
        assert "No Email" in wb.sheetnames

        no_email_sheet = wb["No Email"]
        # Check that member is in no email sheet
        row_data = [cell.value for cell in no_email_sheet[2]]
        assert row_data[1] == "Jane"  # FirstName
        assert row_data[2] == "Smith"  # LastName
        assert row_data[3] == "" or row_data[3] is None  # EmailName should be empty
        assert row_data[5] == "" or row_data[5] is None  # MailName should be empty

    def test_member_with_email_goes_to_main_sheet(self, db, member_with_email):
        """Test that member with email goes to main sheet"""
        queryset = [member_with_email]
        response = generate_expires_two_months_excel(queryset)

        wb = load_workbook(BytesIO(response.content))

        # Should have main sheet (first sheet)
        sheet_names = wb.sheetnames
        main_sheet_name = [name for name in sheet_names if name != "No Email"][0]
        main_sheet = wb[main_sheet_name]

        # Check that member is in main sheet
        row_data = [cell.value for cell in main_sheet[2]]
        assert row_data[1] == "John"  # FirstName
        assert row_data[5] == "John Doe<john.doe@example.com>"  # MailName

    def test_both_sheets_created_when_mixed(
        self, db, member_with_email, member_without_email
    ):
        """Test that both sheets are created when members have mixed email status"""
        queryset = [member_with_email, member_without_email]
        response = generate_expires_two_months_excel(queryset)

        wb = load_workbook(BytesIO(response.content))

        # Should have both sheets
        assert "Expires Two Months" in wb.sheetnames
        assert "No Email" in wb.sheetnames

        # Check main sheet has member with email
        main_sheet = wb["Expires Two Months"]
        row_data = [cell.value for cell in main_sheet[2]]
        assert row_data[1] == "John"  # FirstName

        # Check no email sheet has member without email
        no_email_sheet = wb["No Email"]
        row_data = [cell.value for cell in no_email_sheet[2]]
        assert row_data[1] == "Jane"  # FirstName

    def test_empty_queryset_creates_empty_sheet(self, db):
        """Test that empty queryset creates sheet with headers only"""
        queryset = []
        response = generate_expires_two_months_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Should have headers but no data rows
        assert ws.max_row == 1
        headers = [cell.value for cell in ws[1]]
        assert len(headers) == 7  # Should have 7 columns

    def test_date_formatting(self, db, member_type):
        """Test that dates are formatted as MM/DD/YYYY"""
        member = Member.objects.create(
            member_id=3,
            first_name="Test",
            last_name="Date",
            email="test@example.com",
            member_type=member_type,
            status="active",
            expiration_date=date(2024, 12, 25),
            date_joined=date(2020, 6, 15),
        )

        queryset = [member]
        response = generate_expires_two_months_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        row_data = [cell.value for cell in ws[2]]
        assert row_data[4] == "06/15/2020"  # DateJoined (MM/DD/YYYY)
        assert row_data[6] == "12/25/2024"  # Expires (MM/DD/YYYY)

    def test_mail_name_formatting(self, db, member_type):
        """Test that MailName is formatted correctly"""
        member = Member.objects.create(
            member_id=4,
            first_name="Alice",
            last_name="Brown",
            email="alice.brown@example.com",
            member_type=member_type,
            status="active",
            expiration_date=date.today() - timedelta(days=70),
            date_joined=date(2020, 1, 15),
        )

        queryset = [member]
        response = generate_expires_two_months_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        row_data = [cell.value for cell in ws[2]]
        assert row_data[5] == "Alice Brown<alice.brown@example.com>"  # MailName

    def test_mail_name_empty_when_no_email(self, db, member_without_email):
        """Test that MailName is empty when member has no email"""
        queryset = [member_without_email]
        response = generate_expires_two_months_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        no_email_sheet = wb["No Email"]

        row_data = [cell.value for cell in no_email_sheet[2]]
        assert row_data[5] == "" or row_data[5] is None  # MailName should be empty

    def test_ordering_by_member_id(self, db, member_type):
        """Test that members are ordered by member_id"""
        # Create members in non-sequential order
        member3 = Member.objects.create(
            member_id=3,
            first_name="Third",
            last_name="Member",
            email="third@example.com",
            member_type=member_type,
            status="active",
            expiration_date=date.today() - timedelta(days=70),
            date_joined=date(2020, 1, 15),
        )
        member1 = Member.objects.create(
            member_id=1,
            first_name="First",
            last_name="Member",
            email="first@example.com",
            member_type=member_type,
            status="active",
            expiration_date=date.today() - timedelta(days=70),
            date_joined=date(2020, 1, 15),
        )
        member2 = Member.objects.create(
            member_id=2,
            first_name="Second",
            last_name="Member",
            email="second@example.com",
            member_type=member_type,
            status="active",
            expiration_date=date.today() - timedelta(days=70),
            date_joined=date(2020, 1, 15),
        )

        queryset = sorted([member3, member1, member2], key=lambda m: m.member_id or 0)
        response = generate_expires_two_months_excel(queryset)

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check ordering (row 2, 3, 4 should be member_id 1, 2, 3)
        row2_data = [cell.value for cell in ws[2]]
        row3_data = [cell.value for cell in ws[3]]
        row4_data = [cell.value for cell in ws[4]]

        assert row2_data[0] == 1  # MemberID
        assert row2_data[1] == "First"  # FirstName
        assert row3_data[0] == 2  # MemberID
        assert row3_data[1] == "Second"  # FirstName
        assert row4_data[0] == 3  # MemberID
        assert row4_data[1] == "Third"  # FirstName

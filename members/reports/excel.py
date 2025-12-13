from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
from datetime import date


def generate_newsletter_excel(members_queryset):
    """Generate Excel export of active members for newsletter distribution"""

    # Split members into groups
    members_with_email = []
    members_without_email = []

    for member in members_queryset:
        if member.email and member.email.strip():
            members_with_email.append(member)
        else:
            members_without_email.append(member)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Column headers
    headers = [
        "MemberID",
        "FirstName",
        "LastName",
        "EmailName",
        "DateJoined",
        "Birthdate",
        "Expires",
        "MailName",
        "FullName",
    ]

    # Helper function to format date
    def format_date(d):
        if d:
            return d.strftime("%m/%d/%Y")
        return ""

    # Helper function to write member row
    def write_member_row(ws, member, row_num):
        mail_name = ""
        if member.email and member.email.strip():
            mail_name = f"{member.first_name} {member.last_name}<{member.email}>"

        ws.append(
            [
                member.member_id or "",
                member.first_name,
                member.last_name,
                member.email or "",
                format_date(member.date_joined),
                format_date(member.milestone_date),
                format_date(member.expiration_date),
                mail_name,
                f"{member.first_name} {member.last_name}",
            ]
        )

    # Create numbered sheets for members with emails (99 per sheet)
    if members_with_email:
        sheet_num = 1
        current_sheet = None
        row_count = 0

        for member in members_with_email:
            # Create new sheet if needed (every 99 rows)
            if row_count == 0:
                current_sheet = wb.create_sheet(title=f"Sheet {sheet_num}")
                # Write headers
                current_sheet.append(headers)
                # Make headers bold
                for cell in current_sheet[1]:
                    cell.font = Font(bold=True)
                row_count = 1

            # Write member row
            write_member_row(current_sheet, member, row_count + 1)
            row_count += 1

            # Start new sheet if we've reached 99 members
            if row_count > 99:
                sheet_num += 1
                row_count = 0

    # Create "No Email" sheet
    if members_without_email:
        no_email_sheet = wb.create_sheet(title="No Email")
        # Write headers
        no_email_sheet.append(headers)
        # Make headers bold
        for cell in no_email_sheet[1]:
            cell.font = Font(bold=True)

        # Write all members without emails
        for member in members_without_email:
            write_member_row(no_email_sheet, member, 0)

    # Ensure at least one sheet exists (for empty queryset case)
    if len(wb.sheetnames) == 0:
        empty_sheet = wb.create_sheet(title="Sheet 1")
        empty_sheet.append(headers)
        for cell in empty_sheet[1]:
            cell.font = Font(bold=True)

    # Save to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create HTTP response
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="newsletter_data_{date.today().strftime("%Y_%m_%d")}.xlsx"'
    )

    return response


def generate_new_member_excel(members_queryset):
    """Generate Excel export of new members (active members who joined within date range)"""

    # Split members into groups
    members_with_email = []
    members_without_email = []

    for member in members_queryset:
        if member.email and member.email.strip():
            members_with_email.append(member)
        else:
            members_without_email.append(member)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Column headers (in order as specified - exact formatting for template)
    headers = [
        "MemberID",
        "FirstName",
        "LastName",
        "AddressLong",
        "HomeAddress",
        "HomeCity",
        "HomeState",
        "Zip5",
        "Phone",
        "EmailName",
        "Birthdate",
        "DateJoined",
        "Expires",
        "MailName",
    ]

    # Helper function to format date
    def format_date(d):
        if d:
            return d.strftime("%m/%d/%Y")
        return ""

    # Helper function to extract first 5 digits of zip
    def extract_zip5(zip_str):
        if zip_str:
            # Take first 5 characters (handles "12345" or "12345-6789")
            return zip_str[:5]
        return ""

    # Helper function to create address_long (street[TAB]city, state[TAB]zip)
    def create_address_long(member):
        parts = []
        # Street address (first part)
        if member.home_address:
            parts.append(member.home_address)

        # City and state combined with comma+space (second part)
        city_state_parts = []
        if member.home_city:
            city_state_parts.append(member.home_city)
        if member.home_state:
            city_state_parts.append(member.home_state)
        if city_state_parts:
            # Join city and state with comma+space: "San Jose, CA"
            parts.append(", ".join(city_state_parts))

        # Zip code (third part)
        zip5 = extract_zip5(member.home_zip)
        if zip5:
            parts.append(zip5)

        # Join parts with tab: street[TAB]city, state[TAB]zip
        return "\t".join(parts)

    # Helper function to create mail_name
    def create_mail_name(member):
        if member.email and member.email.strip():
            return f"{member.first_name} {member.last_name}<{member.email}>"
        return ""

    # Helper function to write member row
    def write_member_row(ws, member):
        ws.append(
            [
                member.member_id or "",
                member.first_name,
                member.last_name,
                create_address_long(member),
                member.home_address or "",
                member.home_city or "",
                member.home_state or "",  # Separate HomeState column
                extract_zip5(member.home_zip),  # Separate Zip5 column
                member.home_phone or "",
                member.email or "",
                format_date(member.milestone_date),
                format_date(member.date_joined),
                format_date(member.expiration_date),
                create_mail_name(member),
            ]
        )

    # Create main sheet(s) for members with emails (no 99-row limit)
    if members_with_email:
        main_sheet = wb.create_sheet(title="New Members")
        # Write headers
        main_sheet.append(headers)
        # Make headers bold
        for cell in main_sheet[1]:
            cell.font = Font(bold=True)

        # Write all members with emails
        for member in members_with_email:
            write_member_row(main_sheet, member)

    # Create "no email" sheet if any members lack emails
    if members_without_email:
        no_email_sheet = wb.create_sheet(title="no email")
        # Write headers
        no_email_sheet.append(headers)
        # Make headers bold
        for cell in no_email_sheet[1]:
            cell.font = Font(bold=True)

        # Write all members without emails
        for member in members_without_email:
            write_member_row(no_email_sheet, member)

    # Ensure at least one sheet exists (for empty queryset case)
    if len(wb.sheetnames) == 0:
        empty_sheet = wb.create_sheet(title="New Members")
        empty_sheet.append(headers)
        for cell in empty_sheet[1]:
            cell.font = Font(bold=True)

    # Save to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create HTTP response
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="new_members_{date.today().strftime("%Y_%m_%d")}.xlsx"'
    )

    return response


def generate_milestone_excel(members_queryset):
    """Generate Excel export of active members whose milestone dates fall within date range"""

    # Split members into groups
    members_with_email = []
    members_without_email = []

    for member in members_queryset:
        if member.email and member.email.strip():
            members_with_email.append(member)
        else:
            members_without_email.append(member)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Column headers (in order, 11 columns total)
    headers = [
        "MemberID",
        "FirstName",
        "LastName",
        "Birthdate",
        "DateJoined",
        "Expires",
        "BdayLong",
        "Years",
        "MailName",
        "Jmonth",
        "Jyear",
    ]

    # Helper function to format date
    def format_date(d):
        if d:
            return d.strftime("%m/%d/%Y")
        return ""

    # Helper function to calculate years (current_year - milestone_year)
    def calculate_years(milestone_date):
        if milestone_date:
            current_year = date.today().year
            return current_year - milestone_date.year
        return ""

    # Helper function to get month name from date
    def get_month_name(date_obj):
        if not date_obj:
            return ""
        month_names = [
            "",
            "January",
            "February",
            "March",
            "April",
            "May",
            "June",
            "July",
            "August",
            "September",
            "October",
            "November",
            "December",
        ]
        return month_names[date_obj.month]

    # Helper function to get day of week name
    def get_day_of_week_name(date_obj):
        if not date_obj:
            return ""
        day_names = [
            "Monday",
            "Tuesday",
            "Wednesday",
            "Thursday",
            "Friday",
            "Saturday",
            "Sunday",
        ]
        return day_names[date_obj.weekday()]

    # Helper function to get BdayLong (day of week + date for milestone THIS YEAR)
    def get_bday_long(milestone_date):
        if not milestone_date:
            return ""

        current_year = date.today().year
        milestone_month = milestone_date.month
        milestone_day = milestone_date.day

        # Handle leap year dates (Feb 29) - use Feb 28 in non-leap years
        try:
            milestone_this_year = date(current_year, milestone_month, milestone_day)
        except ValueError:
            # Leap year date in non-leap year - use Feb 28
            milestone_this_year = date(current_year, 2, 28)

        # Get day of week name
        day_name = get_day_of_week_name(milestone_this_year)
        # Get month name
        month_name = get_month_name(milestone_this_year)
        # Format as "Monday, June 15"
        return f"{day_name}, {month_name} {milestone_day}"

    # Helper function to get year from date
    def get_year(date_obj):
        if date_obj:
            return date_obj.year
        return ""

    # Helper function to create mail_name
    def create_mail_name(member):
        if member.email and member.email.strip():
            return f"{member.first_name} {member.last_name}<{member.email}>"
        return ""

    # Helper function to write member row
    def write_member_row(ws, member):
        ws.append(
            [
                member.member_id or "",
                member.first_name,
                member.last_name,
                format_date(member.milestone_date),  # Birthdate
                format_date(member.date_joined),  # DateJoined
                format_date(member.expiration_date),  # Expires
                get_bday_long(member.milestone_date),  # BdayLong
                calculate_years(member.milestone_date),  # Years
                create_mail_name(member),  # MailName
                get_month_name(member.date_joined),  # Jmonth (from date_joined)
                get_year(member.date_joined),  # Jyear (from date_joined)
            ]
        )

    # Create main sheet(s) for members with emails (no row limit)
    if members_with_email:
        main_sheet = wb.create_sheet(title="Milestone Export")
        # Write headers
        main_sheet.append(headers)
        # Make headers bold
        for cell in main_sheet[1]:
            cell.font = Font(bold=True)

        # Write all members with emails
        for member in members_with_email:
            write_member_row(main_sheet, member)

    # Create "no email" sheet if any members lack emails
    if members_without_email:
        no_email_sheet = wb.create_sheet(title="no email")
        # Write headers
        no_email_sheet.append(headers)
        # Make headers bold
        for cell in no_email_sheet[1]:
            cell.font = Font(bold=True)

        # Write all members without emails
        for member in members_without_email:
            write_member_row(no_email_sheet, member)

    # Ensure at least one sheet exists (for empty queryset case)
    if len(wb.sheetnames) == 0:
        empty_sheet = wb.create_sheet(title="Milestone Export")
        empty_sheet.append(headers)
        for cell in empty_sheet[1]:
            cell.font = Font(bold=True)

    # Save to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create HTTP response
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="milestone_export_{date.today().strftime("%Y_%m_%d")}.xlsx"'
    )

    return response
